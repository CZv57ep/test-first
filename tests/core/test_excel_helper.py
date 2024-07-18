import gc
from pathlib import Path

import mergeron.core.excel_helper as xlh
import numpy as np
import pytest
from openpyxl import load_workbook
from pytest import TempPathFactory
from xlsxwriter import Workbook  # type: ignore


@pytest.fixture(scope="session")
def excel_file_path(tmp_path_factory: TempPathFactory) -> Path:
    """Create a path to the test workbook, for writing to and reading from it

    See, https://docs.pytest.org/en/7.1.x/how-to/tmp_path.html
    """
    _fpath = tmp_path_factory.mktemp("excel_test") / "workbook.xlsx"
    return _fpath


def test_excel(excel_file_path: Path) -> None:
    """
    Test functions to create Excel instance and write to worksheet.
    """

    _hdr_val = ("col1", "col2")
    _data_val = np.arange(16).reshape(8, 2)

    _xl_book = Workbook(excel_file_path)
    _xl_sheet = _xl_book.add_worksheet()

    for _col_idx in range(2):
        xlh.scalar_to_sheet(
            _xl_book,
            _xl_sheet,
            0,
            _col_idx,
            _hdr_val[_col_idx],
            (xlh.CFmt.HDR_BORDER, xlh.CFmt.A_RIGHT),
        )

    xlh.array_to_sheet(
        _xl_book,
        _xl_sheet,
        _data_val,
        1,
        0,
        cell_format=(xlh.CFmt.PCT_NUM, xlh.CFmt.A_RIGHT),
        green_bar_flag=True,
    )
    _xl_sheet.set_column(0, 1, 15)
    _xl_book.close()
    del _xl_book

    _xl_book = load_workbook(excel_file_path)
    _xl_sheet = _xl_book[_xl_book.sheetnames[0]]
    _xl_sheet_vals = tuple(_xl_sheet.values)
    _xl_book.close()

    try:
        assert all((  # noqa: S101
            _xl_sheet_vals[0] == _hdr_val,
            np.array_equal(np.array(_xl_sheet_vals[1:]), _data_val),
        ))
    except AssertionError as _aer:
        print(_xl_sheet_vals, _hdr_val, _data_val)
        raise _aer

    gc.collect()
