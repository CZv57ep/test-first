import gc
from collections.abc import Sequence
from pathlib import Path
from typing import Any

import mergeron.core.xlsxw_helper as xlh
import numpy as np
import pytest
from numpy.typing import NDArray
from openpyxl import load_workbook
from pytest import TempPathFactory
from xlsxwriter import Workbook  # type: ignore


@pytest.fixture(scope="session")
def excel_file_path(request: Any, tmp_path_factory: TempPathFactory) -> Path:
    """Create a path to the test workbook, for writing to and reading from it

    See, https://docs.pytest.org/en/7.1.x/how-to/tmp_path.html
    """
    _fpath = tmp_path_factory.mktemp("excel_test")
    return _fpath


@pytest.mark.parametrize(
    ("_data_val", "_ragged_flag", "_cell_format", "_wbk_name"),
    (
        (
            np.arange(16).reshape(8, 2),
            False,
            (xlh.CFmt.PCT_NUM, xlh.CFmt.A_RIGHT),
            "workbook_0.xlsx",
        ),
        (
            [
                (1, 2),
                (3, [4, 5, 6]),
                (5, 6),
                (7, 8),
                (9, 10),
                (11, 12),
                (13, 14),
                (15, 16),
            ],
            True,
            xlh.CFmt.A_RIGHT,
            "workbook_1.xlsx",
        ),
    ),
)
def test_excel(
    _data_val: NDArray[np.int64] | Sequence[Sequence[Any]],
    _ragged_flag: bool,
    _cell_format: Sequence[xlh.CFmt | Sequence[xlh.CFmt]],
    _wbk_name: str,
    excel_file_path: Path,
) -> None:
    """
    Test functions to create Excel instance and write to worksheet.
    """

    _excel_file_path = excel_file_path / _wbk_name
    _hdr_val = ("col1", "col2")

    with Workbook(_excel_file_path) as _xl_book:
        _xl_sheet = _xl_book.add_worksheet()

        for _col_idx, _col_val in enumerate(_hdr_val):
            xlh.scalar_to_sheet(
                _xl_book,
                _xl_sheet,
                0,
                _col_idx,
                _col_val,
                (xlh.CFmt.HDR_BORDER, xlh.CFmt.A_RIGHT),
            )

        xlh.array_to_sheet(
            _xl_book,
            _xl_sheet,
            _data_val,
            1,
            0,
            cell_format=_cell_format,
            green_bar_flag=True,
            ragged_flag=_ragged_flag,
        )
        _xl_sheet.set_column(0, 1, 15)

    _xl_book = load_workbook(_excel_file_path, read_only=True)
    _xl_sheet = _xl_book[_xl_book.sheetnames[0]]
    _xl_sheet_vals = tuple(_xl_sheet.values)
    _xl_book.close()

    _test_data_val = _data_val
    if _ragged_flag:
        _test_data_val = []
        for _row in _data_val:
            _test_row = []
            for _col in _row:
                _test_row.append(
                    repr(_col)
                    if not isinstance(_col, str) and hasattr(_col, "__len__")
                    else _col
                )
            _test_data_val.append(_test_row)

    try:
        assert all((  # noqa: S101
            _xl_sheet_vals[0] == _hdr_val,
            np.array_equal(np.array(_xl_sheet_vals[1:]), _test_data_val),
        ))
    except AssertionError as _aer:
        print(_xl_sheet_vals, _hdr_val, _test_data_val)
        raise _aer

    gc.collect()
